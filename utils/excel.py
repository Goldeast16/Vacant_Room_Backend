from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from zoneinfo import ZoneInfo
from bson import ObjectId
from typing import Any, Dict, List, Set
import json

KST = ZoneInfo("Asia/Seoul")

# 엑셀 헤더 우선순위(존재하는 키만 반영)
PREFERRED_ORDER = [
    "_id", "created_at", "updated_at",
    "name", "email", "phone",
    "category", "message", "anonymous",
    "page_url",
]

def _to_cell_value(v: Any) -> Any:
    if isinstance(v, ObjectId):
        return str(v)
    if isinstance(v, datetime):
        return v.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v if v is not None else ""

def _build_headers(docs: List[Dict[str, Any]]) -> List[str]:
    all_keys: Set[str] = set()
    for d in docs:
        all_keys.update(d.keys())
    ordered = [k for k in PREFERRED_ORDER if k in all_keys]
    remaining = [k for k in sorted(all_keys - set(ordered))]
    return ordered + remaining

def _auto_fit(ws):
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            val = "" if val is None else str(val)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def make_feedback_excel(docs: List[Dict[str, Any]]) -> Workbook:
    normalized = [{k: _to_cell_value(v) for k, v in d.items()} for d in docs]
    headers = _build_headers(normalized)

    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback"

    ws.append(headers)
    for d in normalized:
        ws.append([d.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_fit(ws)

    return wb